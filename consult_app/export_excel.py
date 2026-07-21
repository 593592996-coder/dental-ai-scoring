#!/usr/bin/env python3
"""将cases.py导出为Excel表格，含图片导出功能"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from cases import CASES
import os, shutil

wb = openpyxl.Workbook()
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
cell_font = Font(name='微软雅黑', size=10)
wrap = Alignment(wrap_text=True, vertical='top')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
base = os.path.dirname(os.path.abspath(__file__))
img_dir = os.path.join(base, 'static', 'images')
os.makedirs(img_dir, exist_ok=True)

# ═══════════════════════════════
# Sheet 1: 病例基本信息（含图片导入指引）
# ═══════════════════════════════
ws4 = wb.active
ws4.title = "病例基本信息"

headers4 = ['病例ID', '病例标题（仅显示症状）', '难度', '患者姓名', '性别', '年龄', '职业',
            '主诉', '影像1:图片文件名', '影像1:描述', '影像2:图片文件名', '影像2:描述', '隐藏诊断答案']
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin

# 加使用说明
note_row = len(CASES) + 3
ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=13)
ws4.cell(row=note_row, column=1, value='📌 图片导入说明：将临床照片/X线片命名为如 case001_img1.jpg 放入 static/images/ 文件夹，然后在"影像文件名"列填入文件名即可。系统会自动加载。').font = Font(name='微软雅黑', size=10, color='FFe67e22')
ws4.row_dimensions[note_row].height = 25

row = 2
for cid, case in CASES.items():
    imgs = case.get('images', [])
    img1_name = ''
    img2_name = ''
    for i, img in enumerate(imgs):
        # 尝试查找本地图片
        for ext in ['.png','.jpg','.jpeg','.gif']:
            candidate = os.path.join(img_dir, f'{cid}_img{i+1}{ext}')
            if os.path.exists(candidate):
                if i == 0: img1_name = f'{cid}_img{i+1}{ext}'
                else: img2_name = f'{cid}_img{i+1}{ext}'
                break

    data = [
        cid, case['title'], case.get('difficulty', ''),
        case['patient']['name'], case['patient']['gender'], str(case['patient']['age']),
        case['patient'].get('occupation', ''), case.get('chief_complaint', ''),
        img1_name, imgs[0]['desc'] if len(imgs) > 0 else '',
        img2_name, imgs[1]['desc'] if len(imgs) > 1 else '',
        case.get('diagnosis_answer', ''),
    ]
    for col, val in enumerate(data, 1):
        c = ws4.cell(row=row, column=col, value=val)
        c.font = cell_font; c.alignment = wrap; c.border = thin
    ws4.row_dimensions[row].height = 40
    row += 1

for col, w in enumerate([10,18,10,10,8,6,10,40,20,45,20,45,25], 1):
    ws4.column_dimensions[chr(64+col)].width = w

# ═══════════════════════════════
# Sheet 2: 对话词库
# ═══════════════════════════════
ws1 = wb.create_sheet("对话词库")
headers1 = ['病例ID', '病例标题', '关键词（|分隔）', '患者回答', '所属分类']
for col, h in enumerate(headers1, 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin

row = 2
for cid, case in CASES.items():
    for keywords, answer in case.get('conversation', {}).items():
        ks = keywords.split('|')
        if any(k in keywords for k in ['位置','部位','哪里']): cat = '主诉-部位'
        elif any(k in keywords for k in ['多久','时间','几天','什么时候']): cat = '主诉-时间'
        elif any(k in keywords for k in ['加重','缓解','诱因','情况']): cat = '主诉-诱因'
        elif any(k in keywords for k in ['疼','痛','感觉','性质']): cat = '现病史-疼痛'
        elif any(k in keywords for k in ['自己','自发','夜间','晚上']): cat = '现病史-自发痛'
        elif any(k in keywords for k in ['药','处理','吃过']): cat = '现病史-用药'
        elif any(k in keywords for k in ['病','全身','身体']): cat = '既往史-全身病'
        elif any(k in keywords for k in ['过敏']): cat = '既往史-过敏'
        elif any(k in keywords for k in ['看过','牙科','以前','治']): cat = '既往史-牙科史'
        elif any(k in keywords for k in ['抽烟','喝酒']): cat = '既往史-生活习惯'
        elif any(k in keywords for k in ['职业','工作','学校']): cat = '基本信息'
        elif any(k in keywords for k in ['怎么','原因','受伤','撞']): cat = '主诉-外伤原因'
        elif any(k in keywords for k in ['松','晃']): cat = '现病史-松动'
        else: cat = '其他'
        for col, val in enumerate([cid, case['title'], keywords, answer, cat], 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = cell_font; c.alignment = wrap; c.border = thin
        row += 1

for col, w in enumerate([10,16,35,60,16], 1):
    ws1.column_dimensions[chr(64+col)].width = w

# ═══════════════════════════════
# Sheet 3: 检查标准库
# ═══════════════════════════════
ws2 = wb.create_sheet("检查标准库")
headers2 = ['病例ID', '病例标题', '检查项目', '正确顺序', '标准器械（逗号分隔）', '标准操作技术', '关键操作要点', '临床发现']
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin

row = 2
for cid, case in CASES.items():
    exam = case.get('examination', {})
    correct_seq = exam.get('correct_sequence', [])
    for item_name, item_data in exam.get('items', {}).items():
        seq_num = correct_seq.index(item_name) + 1 if item_name in correct_seq else 0
        for col, val in enumerate([
            cid, case['title'], item_name, f'第{seq_num}步',
            ', '.join(item_data.get('instruments', [])), item_data.get('technique', ''),
            item_data.get('key_points', ''), item_data.get('finding', ''),
        ], 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = cell_font; c.alignment = wrap; c.border = thin
        row += 1

for col, w in enumerate([10,16,20,10,30,50,30,55], 1):
    ws2.column_dimensions[chr(64+col)].width = w

# ═══════════════════════════════
# Sheet 4: 病历标准库
# ═══════════════════════════════
ws3 = wb.create_sheet("病历标准库")
headers3 = ['病例ID', '病例标题', '标准诊断', '可接受诊断（逗号分隔）', '鉴别诊断（逗号分隔）',
            '治疗计划（每行一项）', '处置步骤', '医嘱（每行一项）']
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin

row = 2
for cid, case in CASES.items():
    rec = case.get('medical_record', {})
    for col, val in enumerate([
        cid, case['title'], rec.get('diagnosis', ''),
        ', '.join(rec.get('acceptable_diag', [])), ', '.join(rec.get('differential', [])),
        '\n'.join(rec.get('treatment', [])), rec.get('procedure', ''),
        '\n'.join(rec.get('orders', [])),
    ], 1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = cell_font; c.alignment = wrap; c.border = thin
    row += 1

for col, w in enumerate([10,16,30,30,30,40,45,35], 1):
    ws3.column_dimensions[chr(64+col)].width = w

# 冻结首行
for ws in [ws1, ws2, ws3, ws4]:
    ws.freeze_panes = 'A2'
    if ws.max_row > 1: ws.auto_filter.ref = ws.dimensions

output = os.path.join(base, '词库管理.xlsx')
wb.save(output)
print(f'✅ 已导出: {output}')
print(f'   Sheet 1: 病例基本信息（含图片文件名列）')
print(f'   Sheet 2: 对话词库 ({ws1.max_row-1}条)')
print(f'   Sheet 3: 检查标准库 ({ws2.max_row-1}条)')
print(f'   Sheet 4: 病历标准库 ({ws3.max_row-1}条)')
print(f'\n📌 图片存放目录: {img_dir}/')
print(f'   命名格式: 病例ID_img1.jpg / 病例ID_img2.jpg')
print(f'   示例: case001_img1.jpg, case001_img2.jpg')
