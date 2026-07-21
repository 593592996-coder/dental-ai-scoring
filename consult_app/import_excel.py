#!/usr/bin/env python3
"""
从Excel导入数据，生成cases.py
支持图片：在Excel"病例基本信息"Sheet的"影像文件名"列填入文件名
图片放在 static/images/ 文件夹下
"""
import openpyxl, sys, os, shutil

def import_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    base = os.path.dirname(os.path.abspath(excel_path))
    img_dir = os.path.join(base, 'static', 'images')
    os.makedirs(img_dir, exist_ok=True)

    # ═══ Sheet 1: 病例基本信息 ═══
    cases_data = {}
    ws4 = wb['病例基本信息']
    for row in ws4.iter_rows(min_row=2, values_only=True):
        vals = list(row)[:14]  # 扩展列数
        if not vals[0]: break  # 空行结束
        cid, title, diff, pname, pgender, page, pocc, chief, img1_file, img1_desc, img2_file, img2_desc, answer = vals[:13]

        images = []
        # 图片1
        if img1_file and img1_file.strip():
            img_path = os.path.join(img_dir, img1_file.strip())
            if os.path.exists(img_path):
                images.append({'label': '影像1', 'desc': str(img1_desc or ''), 'file': str(img1_file).strip()})
            else:
                print(f'  ⚠️ {cid}: 图片 {img1_file} 不在 {img_dir}/ 中，跳过')
                images.append({'label': '影像1', 'desc': str(img1_desc or ''), 'file': ''})
        else:
            images.append({'label': '影像1', 'desc': str(img1_desc or ''), 'file': ''}) if img1_desc else None

        # 图片2
        if img2_file and img2_file.strip():
            img_path2 = os.path.join(img_dir, img2_file.strip())
            if os.path.exists(img_path2):
                images.append({'label': '影像2', 'desc': str(img2_desc or ''), 'file': str(img2_file).strip()})
            else:
                print(f'  ⚠️ {cid}: 图片 {img2_file} 不在 {img_dir}/ 中，跳过')
                images.append({'label': '影像2', 'desc': str(img2_desc or ''), 'file': ''})
        else:
            if img2_desc: images.append({'label': '影像2', 'desc': str(img2_desc), 'file': ''})

        # 清理空desc项
        images = [img for img in images if img['desc'].strip()]

        level = 1
        if diff and '★★★' in str(diff): level = 3
        elif diff and '★★' in str(diff): level = 2

        cases_data[cid] = {
            'id': cid, 'title': title or '', 'difficulty': diff or '★', 'level': level,
            'patient': {'name': pname or '', 'gender': pgender or '', 'age': int(page) if page else 0, 'occupation': str(pocc or '')},
            'chief_complaint': str(chief or ''),
            'images': images,
            'diagnosis_answer': str(answer or ''),
            'conversation': {},
            'examination': {'correct_sequence': [], 'items': {}},
            'medical_record': {'diagnosis': '', 'acceptable_diag': [], 'differential': [], 'treatment': [], 'procedure': '', 'orders': []},
        }
        print(f'  ✅ {cid}: {title} (图片{len(images)}张)')

    # ═══ Sheet 2: 对话词库 ═══
    if '对话词库' in wb.sheetnames:
        ws1 = wb['对话词库']
        for row in ws1.iter_rows(min_row=2, values_only=True):
            vals = list(row)[:5]
            if not vals[0] or not vals[2]: continue
            cid, _, keywords, answer, _ = vals
            if cid in cases_data and keywords:
                cases_data[cid]['conversation'][str(keywords)] = str(answer or '')

    # ═══ Sheet 3: 检查标准库 ═══
    if '检查标准库' in wb.sheetnames:
        ws2 = wb['检查标准库']
        for row in ws2.iter_rows(min_row=2, values_only=True):
            vals = list(row)[:8]
            if not vals[0] or not vals[2]: continue
            cid, _, item_name, seq, instruments, technique, key_points, finding = vals
            if cid in cases_data and item_name:
                instr_list = [i.strip() for i in str(instruments).split(',') if i.strip()] if instruments else []
                cases_data[cid]['examination']['items'][str(item_name)] = {
                    'instruments': instr_list, 'technique': str(technique or ''),
                    'key_points': str(key_points or ''), 'finding': str(finding or ''),
                }
                if str(item_name) not in cases_data[cid]['examination']['correct_sequence']:
                    cases_data[cid]['examination']['correct_sequence'].append(str(item_name))

    # ═══ Sheet 4: 病历标准库 ═══
    if '病历标准库' in wb.sheetnames:
        ws3 = wb['病历标准库']
        for row in ws3.iter_rows(min_row=2, values_only=True):
            vals = list(row)[:8]
            if not vals[0]: continue
            cid, _, diagnosis, acceptable, differential, treatment, procedure, orders = vals
            if cid in cases_data:
                rec = cases_data[cid]['medical_record']
                rec['diagnosis'] = str(diagnosis or '')
                rec['acceptable_diag'] = [a.strip() for a in str(acceptable).split(',') if a.strip()] if acceptable else []
                rec['differential'] = [d.strip() for d in str(differential).split(',') if d.strip()] if differential else []
                rec['treatment'] = [t.strip() for t in str(treatment).split('\n') if t.strip()] if treatment else []
                rec['procedure'] = str(procedure or '')
                rec['orders'] = [o.strip() for o in str(orders).split('\n') if o.strip()] if orders else []

    # ═══ 生成 cases.py ═══
    lines = ['"""口腔AI模拟诊疗系统 — 病例数据库（由Excel自动生成）"""\n']
    lines.append('CASES = {\n')

    for cid, c in cases_data.items():
        lines.append(f'    "{cid}": {{\n')
        lines.append(f'        "id": "{c["id"]}",\n')
        lines.append(f'        "title": "{c["title"]}",\n')
        lines.append(f'        "difficulty": "{c["difficulty"]}",\n')
        lines.append(f'        "level": {c["level"]},\n')
        lines.append(f'        "patient": {{"name": "{c["patient"]["name"]}", "gender": "{c["patient"]["gender"]}", "age": {c["patient"]["age"]}, "occupation": "{c["patient"]["occupation"]}"}},\n')
        lines.append(f'        "chief_complaint": """{c["chief_complaint"]}""",\n')
        # images
        lines.append(f'        "images": [\n')
        for img in c['images']:
            if img.get('file'):
                lines.append(f'            {{"label": "{img["label"]}", "desc": """{img["desc"]}""", "file": "{img["file"]}"}},\n')
            else:
                lines.append(f'            {{"label": "{img["label"]}", "desc": """{img["desc"]}"""}},\n')
        lines.append(f'        ],\n')
        lines.append(f'        "diagnosis_answer": """{c["diagnosis_answer"]}""",\n')
        # conversation
        lines.append(f'        "conversation": {{\n')
        for kw, ans in c['conversation'].items():
            lines.append(f'            "{kw}": """{ans}""",\n')
        lines.append(f'        }},\n')
        # examination
        lines.append(f'        "examination": {{\n')
        lines.append(f'            "correct_sequence": {c["examination"]["correct_sequence"]},\n')
        lines.append(f'            "items": {{\n')
        for item_name, item in c['examination']['items'].items():
            lines.append(f'                "{item_name}": {{\n')
            lines.append(f'                    "instruments": {item["instruments"]},\n')
            lines.append(f'                    "technique": """{item["technique"]}""",\n')
            lines.append(f'                    "key_points": """{item["key_points"]}""",\n')
            lines.append(f'                    "finding": """{item["finding"]}""",\n')
            lines.append(f'                }},\n')
        lines.append(f'            }},\n')
        lines.append(f'        }},\n')
        # medical_record
        rec = c['medical_record']
        lines.append(f'        "medical_record": {{\n')
        lines.append(f'            "diagnosis": """{rec["diagnosis"]}""",\n')
        lines.append(f'            "acceptable_diag": {rec["acceptable_diag"]},\n')
        lines.append(f'            "differential": {rec["differential"]},\n')
        lines.append(f'            "treatment": {rec["treatment"]},\n')
        lines.append(f'            "procedure": """{rec["procedure"]}""",\n')
        lines.append(f'            "orders": {rec["orders"]},\n')
        lines.append(f'        }},\n')
        lines.append(f'    }},\n')

    lines.append('}\n')

    output = os.path.join(base, 'cases.py')
    with open(output, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'\n✅ 已生成: {output} ({len(cases_data)}个病例)')
    print(f'   图片目录: {img_dir}/')
    return output


if __name__ == '__main__':
    if len(sys.argv) > 1:
        import_excel(sys.argv[1])
    else:
        default = '/Users/ouyangjunhan/Desktop/AI问诊系统/词库管理.xlsx'
        if os.path.exists(default):
            import_excel(default)
        else:
            print('用法: python import_excel.py 词库管理.xlsx')
