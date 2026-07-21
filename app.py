#!/usr/bin/env python3
"""
II类洞AI评分系统 — Flask Web应用
支持：学生上传照片 → AI分析 → 评分报告 → 教师dashboard
"""

import os
import json
import time
import uuid
from datetime import datetime
from pathlib import Path

from flask import (Flask, render_template, request, redirect,
                   url_for, jsonify, send_from_directory, session, send_file)
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import os, json, uuid, time
from datetime import datetime

from scoring_engine_v2 import II类洞评分引擎V2, SCORING_CONFIG
from scoring_endodontic import 开髓术评分引擎, SCORING_CONFIG as ENDO_CONFIG
from scoring_xray import 根管X光片评估引擎, SCORING_CONFIG as XRAY_CONFIG
from cases_consult import CASES as CONSULT_CASES

# 问诊系统路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONSULT_LOG_DIR = os.path.join(BASE_DIR, 'consult_logs')
CONSULT_IMG_DIR = os.path.join(BASE_DIR, 'static', 'images')
os.makedirs(CONSULT_LOG_DIR, exist_ok=True)
os.makedirs(CONSULT_IMG_DIR, exist_ok=True)

# ── App Setup ──
BASE_DIR = Path(__file__).parent.absolute()
UPLOAD_FOLDER = BASE_DIR / 'uploads'
REPORT_FOLDER = BASE_DIR / 'reports'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'bmp', 'tiff'}

app = Flask(__name__)
app.secret_key = 'kouqiang_neike_2026_ai_scoring'
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

UPLOAD_FOLDER.mkdir(exist_ok=True)
REPORT_FOLDER.mkdir(exist_ok=True)

# 全局评分历史（简化版，生产环境应使用数据库）
scoring_history = []

# 启动时加载历史报告
for f in sorted(REPORT_FOLDER.glob('*.json')):
    try:
        with open(f, 'r', encoding='utf-8') as fh:
            scoring_history.append(json.load(fh))
    except:
        pass

engine = II类洞评分引擎V2()
endo_engine = 开髓术评分引擎()
xray_engine = 根管X光片评估引擎()


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Routes ──
@app.route('/')
def index():
    """统一入口 — 三个模块选择"""
    return render_template('home.html')


@app.route('/class2')
def class2_index():
    """II类洞评分 — 学生端"""
    return render_template('index.html', config=SCORING_CONFIG)


@app.route('/endo')
def endo_index():
    """开髓洞形评分 — 学生端"""
    return render_template('endo.html', config=ENDO_CONFIG)


@app.route('/xray')
def xray_index():
    """根管X光片评估 — 学生端"""
    return render_template('xray.html', config=XRAY_CONFIG)


# ═══════════════════════════════
# 模拟诊疗问诊系统路由（三阶段：自由对话问诊→检查器械技术→病历书写）
# ═══════════════════════════════
@app.route('/consult')
def consult_index():
    return render_template('consult.html', cases=CONSULT_CASES)


@app.route('/api/consult/cases')
def consult_list_cases():
    return jsonify({cid: {'id': cid, 'title': c['title'], 'difficulty': c['difficulty'],
                          'patient': c['patient']} for cid, c in CONSULT_CASES.items()})


@app.route('/api/consult/start/<case_id>', methods=['POST'])
def consult_start(case_id):
    if case_id not in CONSULT_CASES:
        return jsonify({'error': '病例不存在'}), 404
    c = CONSULT_CASES[case_id]
    session.clear()
    session['consult_case_id'] = case_id
    session['consult_phase'] = 1
    session['consult_history_log'] = []
    session['consult_exam_sequence'] = []
    session['consult_exam_details'] = {}
    session['consult_exam_score'] = 0
    session['consult_consult_id'] = uuid.uuid4().hex[:8]
    session['consult_chat_log'] = []
    # 给图片加上完整URL路径
    images_with_url = []
    for img in c.get('images', []):
        img_copy = dict(img)
        if img_copy.get('file'):
            img_copy['url'] = f'/static/images/{img_copy["file"]}'
        images_with_url.append(img_copy)
    return jsonify({
        'chief_complaint': c['chief_complaint'],
        'images': images_with_url,
        'patient': c['patient'],
        'case_title': c['title']
    })


# ══════ 阶段1: 自由对话问诊 ══════
@app.route('/api/consult/chat', methods=['POST'])
def consult_chat():
    case_id = session.get('consult_case_id')
    if not case_id:
        return jsonify({'error': '请先选择病例'}), 400
    c = CONSULT_CASES[case_id]
    data = request.get_json()
    student_input = data.get('message', '').strip()
    if not student_input:
        return jsonify({'reply': '请描述你想了解的情况。', 'coverage': len(session.get('consult_history_log', []))})
    # 关键词匹配（最佳匹配策略 — 双向子串匹配）
    conv = c['conversation']
    best_answer = None
    best_keywords = ''
    best_score = 0
    for keywords, answer in conv.items():
        ks = keywords.split('|')
        score = 0
        for k in ks:
            if k in student_input or student_input in k:
                score += 1 + len(k) * 0.1
        if score > best_score:
            best_score = score
            best_answer = answer
            best_keywords = keywords
    if best_answer:
        reply = best_answer
    else:
        # 记录未匹配问题
        unmatched_file = os.path.join(CONSULT_LOG_DIR, 'unmatched_questions.json')
        try:
            unmatched = []
            if os.path.exists(unmatched_file):
                with open(unmatched_file, 'r', encoding='utf-8') as f:
                    unmatched = json.load(f)
            unmatched.append({
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'question': student_input,
                'case_id': case_id,
                'case_title': c['title']
            })
            with open(unmatched_file, 'w', encoding='utf-8') as f:
                json.dump(unmatched[-100:], f, ensure_ascii=False, indent=2)
        except:
            pass
        reply = '我记不清了，你换个方式问问？'
    log = session.get('consult_history_log', [])
    log.append({'student': student_input, 'patient': reply})
    session['consult_history_log'] = log
    # 计算问诊覆盖度
    covered = sum(1 for kw in conv if any(
        k in ''.join(s['student'] for s in log) for k in kw.split('|')))
    total_kw = len(conv)
    coverage = min(100, int(covered / total_kw * 100)) if total_kw > 0 else 0
    # 记录详细日志
    log_entry = {
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'student': student_input,
        'patient': reply,
        'matched_keywords': [k for k in best_keywords.split('|') if k in student_input] if best_answer else [],
        'all_keywords': best_keywords if best_answer else '',
        'coverage': coverage,
        'case_id': case_id,
    }
    log_data = session.get('consult_chat_log', [])
    log_data.append(log_entry)
    session['consult_chat_log'] = log_data
    # 持久化到文件
    consult_log_file = os.path.join(CONSULT_LOG_DIR, f'{session.get("consult_consult_id", "unknown")}.json')
    try:
        with open(consult_log_file, 'w', encoding='utf-8') as f:
            json.dump({
                'case_id': case_id,
                'case_title': c['title'],
                'log': log_data
            }, f, ensure_ascii=False, indent=2)
    except:
        pass
    return jsonify({
        'reply': reply,
        'coverage': coverage,
        'matched': best_keywords.split('|') if best_answer else []
    })


@app.route('/api/consult/phase1/done', methods=['POST'])
def consult_phase1_done():
    session['consult_phase'] = 2
    case_id = session.get('consult_case_id')
    c = CONSULT_CASES[case_id]
    conv = c['conversation']
    log = session.get('consult_history_log', [])
    # 统计每个关键词组是否被覆盖
    coverage_detail = []
    total_kw_groups = len(conv)
    covered_groups = 0
    for keywords in conv:
        ks = keywords.split('|')
        asked = any(any(k in s['student'] for k in ks) for s in log)
        if asked:
            covered_groups += 1
            coverage_detail.append({'keywords': keywords, 'covered': True, 'status': 'good'})
        else:
            coverage_detail.append({'keywords': keywords, 'covered': False, 'status': 'missed'})
    coverage_pct = int(covered_groups / total_kw_groups * 100) if total_kw_groups > 0 else 0
    # 按类别分组统计
    categories = {
        '主诉': ['位置', '部位', '哪里', '多久', '时间', '几天', '加重', '诱因', '缓解', '情况'],
        '现病史': ['疼', '痛', '性质', '感觉', '自发', '夜间', '自己', '药', '处理', '吃过'],
        '既往史': ['病', '全身', '身体', '过敏', '看过', '牙科', '以前', '治'],
        '生活习惯': ['抽烟', '喝酒', '吸烟', '职业', '工作'],
    }
    cat_scores = {}
    cat_missed = {}
    for cat, kws in categories.items():
        covered = sum(1 for kw_g in conv if any(
            any(k in s['student'] for s in log) for k in kw_g.split('|') if any(ck in k for ck in kws)))
        total = sum(1 for kw_g in conv if any(
            any(ck in k for ck in kws) for k in kw_g.split('|')))
        cat_scores[cat] = {'covered': covered, 'total': total, 'pct': int(covered / max(total, 1) * 100)}
        if total > 0 and covered < total:
            cat_missed[cat] = [
                kw_g for kw_g in conv
                if any(any(ck in k for ck in kws) for k in kw_g.split('|'))
                and not any(any(k in s['student'] for s in log) for k in kw_g.split('|'))
            ]
    score = int(coverage_pct * 1.0)
    suggestions = []
    for cat, missed_kws in cat_missed.items():
        if missed_kws:
            suggestions.append(f'{cat}方面遗漏: {", ".join(missed_kws[:2])}')
    return jsonify({
        'score': score, 'coverage_pct': coverage_pct,
        'covered_groups': covered_groups, 'total_groups': total_kw_groups,
        'coverage_detail': coverage_detail,
        'cat_scores': cat_scores,
        'suggestions': suggestions[:5],
        'message': '病史采集完成。请根据以上信息，设计口腔检查方案。',
        'exam_items': list(c['examination']['items'].keys()),
    })


# ══════ 阶段2: 口腔检查（两步：顺序→器械技术） ══════
@app.route('/api/consult/exam/submit_sequence', methods=['POST'])
def consult_submit_sequence():
    case_id = session.get('consult_case_id')
    if not case_id: return jsonify({'error': ''}), 400
    c = CONSULT_CASES[case_id]
    data = request.get_json()
    student_seq = data.get('sequence', [])
    correct = c['examination']['correct_sequence']
    session['consult_exam_sequence'] = student_seq
    errors = []
    for i, item in enumerate(student_seq):
        if item in correct:
            expected_idx = correct.index(item)
            if abs(i - expected_idx) > 1:
                errors.append(f'"{item}"建议在第{expected_idx + 1}步，你放在第{i + 1}步')
    for ci in correct:
        if ci not in student_seq:
            errors.append(f'缺少：{ci}')
    seq_score = max(0, 10 - len(errors) * 2)
    return jsonify({'errors': errors, 'score': seq_score, 'correct': correct})


@app.route('/api/consult/exam/submit_item', methods=['POST'])
def consult_submit_exam_item():
    case_id = session.get('consult_case_id')
    if not case_id: return jsonify({'error': ''}), 400
    c = CONSULT_CASES[case_id]
    data = request.get_json()
    item = data.get('item', '')
    instruments = data.get('instruments', '')
    technique = data.get('technique', '')
    key_points_input = data.get('key_points', '')
    if item not in c['examination']['items']:
        return jsonify({'error': '检查项目不存在'})
    exam_item = c['examination']['items'][item]
    correct_instruments = set(''.join(exam_item['instruments']).replace('（', '').replace('）', ''))
    student_instruments = set(instruments.replace('（', '').replace('）', ''))
    instr_overlap = len(correct_instruments & student_instruments) / max(len(correct_instruments), 1)
    instr_score = min(10, int(instr_overlap * 10))
    correct_kps = set(exam_item.get('key_points', '').replace('；', ';').split(';'))
    student_kps = set(key_points_input.replace('；', ';').replace('，', ',').split(','))
    kp_overlap = len(correct_kps & student_kps) / max(len(correct_kps), 1) if correct_kps else 0.5
    tech_score = min(10, int(kp_overlap * 10))
    desc_score = min(10, 5 + int(len(set(student_instruments)) / max(len(correct_instruments), 1) * 5))
    total = instr_score + tech_score + desc_score
    session['consult_exam_score'] = session.get('consult_exam_score', 0) + total
    details = session.get('consult_exam_details', {})
    details[item] = {'instruments': instruments, 'technique': technique,
                     'key_points': key_points_input, 'score': total}
    session['consult_exam_details'] = details
    return jsonify({
        'finding': exam_item['finding'],
        'correct_instruments': exam_item['instruments'],
        'correct_technique': exam_item['technique'],
        'correct_key_points': exam_item['key_points'],
        'scores': {'instruments': instr_score, 'technique': tech_score,
                   'description': desc_score, 'total': total}
    })


@app.route('/api/consult/phase2/done', methods=['POST'])
def consult_phase2_done():
    session['consult_phase'] = 3
    case_id = session.get('consult_case_id')
    c = CONSULT_CASES[case_id]
    student_seq = session.get('consult_exam_sequence', [])
    correct_seq = c['examination']['correct_sequence']
    details = session.get('consult_exam_details', {})
    total_items = len(correct_seq)
    seq_errors = []
    for i, item in enumerate(student_seq):
        if item in correct_seq:
            expected = correct_seq.index(item)
            if abs(i - expected) > 1:
                seq_errors.append(f'"{item}"应在第{expected + 1}步')
    missing = [ci for ci in correct_seq if ci not in student_seq]
    seq_score = max(0, 15 - len(seq_errors) * 3 - len(missing) * 5)
    if missing: seq_errors.append(f'缺少: {", ".join(missing)}')
    exam_score = session.get('consult_exam_score', 0)
    max_per_item = 30
    exam_pct = min(100, int(exam_score / (total_items * max_per_item) * 100)) if total_items > 0 else 0
    total = int(seq_score * 0.3 + exam_pct * 0.7)
    suggestions = []
    if missing: suggestions.append(f'检查项目遗漏: {", ".join(missing)}')
    if seq_errors: suggestions.extend(seq_errors[:3])
    if exam_pct < 60: suggestions.append('器械选择或技术描述需加强，请参照标准答案对比学习')
    return jsonify({
        'scores': {'sequence': seq_score, 'technique': exam_pct, 'total': total},
        'completed': f'{len(details)}/{total_items}',
        'errors': seq_errors,
        'suggestions': suggestions,
        'key_findings': [
            f"{item}: {c['examination']['items'].get(item, {}).get('finding', '')[:80]}..."
            for item in student_seq[:5] if item in c['examination']['items']
        ],
    })


# ══════ 阶段3: 病历书写 ══════
@app.route('/api/consult/record/submit', methods=['POST'])
def consult_submit_record():
    case_id = session.get('consult_case_id')
    if not case_id: return jsonify({'error': ''}), 400
    c = CONSULT_CASES[case_id]
    data = request.get_json()
    rec = c['medical_record']
    student_diag = data.get('diagnosis', '').strip()
    correct = rec['diagnosis']
    acceptable = rec.get('acceptable_diag', [])
    if student_diag == correct:
        diag_score = 100
    elif any(a in student_diag or student_diag in a for a in acceptable):
        diag_score = 75
    else:
        kw_s = set(student_diag.replace('，', ',').replace(' ', '').split(','))
        kw_c = set(correct.replace('，', ',').replace(' ', '').split(','))
        overlap = len(kw_s & kw_c) / max(len(kw_c), 1)
        diag_score = int(max(15, overlap * 70))
    student_diff = data.get('differential', '')
    diff_correct = rec.get('differential', [])
    diff_score = sum(33 for d in diff_correct if d in student_diff) if diff_correct else 80
    diff_score = min(100, diff_score)
    student_tx = data.get('treatment_plan', '')
    tx_score = 60
    for kw, pts in {'根管': 20, '开髓': 10, '充填': 10, '修复': 10, '全冠': 10,
                    '活髓': 15, '盖髓': 15, '卫生宣教': 5, '复查': 5, '戒烟': 5}.items():
        if kw in student_tx: tx_score += pts
    tx_score = min(100, tx_score)
    exam_s = session.get('consult_exam_score', 0)
    exam_pct = min(100, exam_s)
    total = int(diag_score * 0.30 + diff_score * 0.15 + tx_score * 0.20 + exam_pct * 0.20 + 80 * 0.15)
    # 生成病历
    p = c['patient']
    record = f"""════════════════════════════════════
           口腔门诊病历
════════════════════════════════════
姓名: {p['name']}  性别: {p['gender']}  年龄: {p['age']}岁

【主诉】{c['chief_complaint']}

【现病史】(学生问诊采集)

【既往史】(学生问诊采集)

【口腔检查】
{(chr(10)).join(f'{item}: {c["examination"]["items"][item]["finding"]}' for item in c['examination']['correct_sequence'] if item in c['examination']['items'])}

【影像学检查】
{(chr(10)).join(f'{img["label"]}: {img["desc"]}' for img in c['images'])}

【诊断】{correct}
【鉴别诊断】{', '.join(diff_correct) if diff_correct else ''}

【治疗计划】
{(chr(10)).join('  ' + t for t in rec['treatment'])}

【处置】
{rec['procedure']}

【医嘱】
{(chr(10)).join('  · ' + o for o in rec['orders'])}

医师: __________  日期: __________"""
    return jsonify({
        'scores': {'diagnosis': diag_score, 'differential': diff_score, 'treatment': tx_score,
                   'examination': exam_pct, 'total': total},
        'correct': {'diagnosis': correct, 'differential': diff_correct, 'treatment': rec['treatment'],
                    'procedure': rec['procedure'], 'orders': rec['orders']},
        'record': record
    })


# ══════ 教师后台 ══════
@app.route('/api/consult/teacher/save', methods=['POST'])
def consult_save_wordbank():
    """在线编辑词库 — 教师新增/修改关键词"""
    data = request.get_json()
    case_id = data.get('case_id', '')
    keywords = data.get('keywords', '').strip()
    answer = data.get('answer', '').strip()
    action = data.get('action', 'add')
    if not case_id or case_id not in CONSULT_CASES:
        return jsonify({'error': '病例不存在'}), 400
    if not keywords or not answer:
        return jsonify({'error': '关键词和回答不能为空'}), 400
    c = CONSULT_CASES[case_id]
    if action == 'delete':
        if keywords in c['conversation']:
            del c['conversation'][keywords]
        message = f'已删除关键词: {keywords}'
    else:
        existed = keywords in c.get('conversation', {})
        c['conversation'][keywords] = answer
        message = f'已{"更新" if existed else "新增"}关键词: {keywords}'
    consult_save_cases_to_file()
    return jsonify({'success': True, 'message': message})


def consult_save_cases_to_file():
    """将当前CONSULT_CASES保存到cases_consult.py"""
    path = os.path.join(BASE_DIR, 'cases_consult.py')
    lines = ['"""口腔AI模拟诊疗系统 — 病例数据库（可由教师在线编辑）"""\n']
    lines.append('CASES = {\n')
    for cid, c in CONSULT_CASES.items():
        lines.append(f'    "{cid}": {{\n')
        lines.append(f'        "id": "{c["id"]}",\n')
        lines.append(f'        "title": "{c["title"]}",\n')
        lines.append(f'        "difficulty": "{c["difficulty"]}",\n')
        lines.append(f'        "level": {c["level"]},\n')
        p = c['patient']
        lines.append(f'        "patient": {{"name": "{p["name"]}", "gender": "{p["gender"]}", '
                     f'"age": {p["age"]}, "occupation": "{p.get("occupation", "")}"}},\n')
        lines.append(f'        "chief_complaint": """{c["chief_complaint"]}""",\n')
        imgs = c.get('images', [])
        lines.append(f'        "images": [\n')
        for img in imgs:
            if img.get('file'):
                lines.append(f'            {{"label": "{img["label"]}", '
                             f'"desc": """{img["desc"]}""", "file": "{img["file"]}"}},\n')
            else:
                lines.append(f'            {{"label": "{img["label"]}", "desc": """{img["desc"]}"""}},\n')
        lines.append(f'        ],\n')
        lines.append(f'        "diagnosis_answer": """{c.get("diagnosis_answer", "")}""",\n')
        conv = c.get('conversation', {})
        lines.append(f'        "conversation": {{\n')
        for kw, ans in conv.items():
            lines.append(f'            "{kw}": """{ans}""",\n')
        lines.append(f'        }},\n')
        exam = c.get('examination', {})
        lines.append(f'        "examination": {{\n')
        lines.append(f'            "correct_sequence": {exam.get("correct_sequence", [])},\n')
        lines.append(f'            "items": {{\n')
        for item_name, item in exam.get('items', {}).items():
            lines.append(f'                "{item_name}": {{\n')
            lines.append(f'                    "instruments": {item.get("instruments", [])},\n')
            lines.append(f'                    "technique": """{item.get("technique", "")}""",\n')
            lines.append(f'                    "key_points": """{item.get("key_points", "")}""",\n')
            lines.append(f'                    "finding": """{item.get("finding", "")}""",\n')
            lines.append(f'                }},\n')
        lines.append(f'            }},\n')
        lines.append(f'        }},\n')
        rec = c.get('medical_record', {})
        lines.append(f'        "medical_record": {{\n')
        lines.append(f'            "diagnosis": """{rec.get("diagnosis", "")}""",\n')
        lines.append(f'            "acceptable_diag": {rec.get("acceptable_diag", [])},\n')
        lines.append(f'            "differential": {rec.get("differential", [])},\n')
        lines.append(f'            "treatment": {rec.get("treatment", [])},\n')
        lines.append(f'            "procedure": """{rec.get("procedure", "")}""",\n')
        lines.append(f'            "orders": {rec.get("orders", [])},\n')
        lines.append(f'        }},\n')
        lines.append(f'    }},\n')
    lines.append('}\n')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


@app.route('/api/consult/teacher/wordbank/<case_id>')
def consult_get_wordbank(case_id):
    if case_id not in CONSULT_CASES:
        return jsonify({}), 404
    c = CONSULT_CASES[case_id]
    return jsonify({
        'case_id': case_id,
        'case_title': c['title'],
        'conversation': c.get('conversation', {}),
        'exam_items': list(c.get('examination', {}).get('items', {}).keys()),
    })


@app.route('/api/consult/teacher/unmatched')
def consult_get_unmatched():
    unmatched_file = os.path.join(CONSULT_LOG_DIR, 'unmatched_questions.json')
    if os.path.exists(unmatched_file):
        with open(unmatched_file, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    return jsonify([])


@app.route('/consult/teacher')
def consult_teacher_dashboard():
    all_logs = []
    case_stats = {}
    for fname in os.listdir(CONSULT_LOG_DIR):
        if fname.endswith('.json') and fname != 'unmatched_questions.json':
            try:
                with open(os.path.join(CONSULT_LOG_DIR, fname), 'r', encoding='utf-8') as f:
                    data = json.load(f)
                data['consult_id'] = fname.replace('.json', '')
                all_logs.append(data)
                cid = data.get('case_id', 'unknown')
                if cid not in case_stats:
                    case_stats[cid] = {'count': 0, 'title': data.get('case_title', ''),
                                       'total_coverage': 0, 'sessions': 0}
                case_stats[cid]['count'] += 1
                case_stats[cid]['sessions'] += 1
                if data.get('log'):
                    last = data['log'][-1]
                    case_stats[cid]['total_coverage'] += last.get('coverage', 0)
            except:
                pass
    for cid in case_stats:
        if case_stats[cid]['sessions'] > 0:
            case_stats[cid]['avg_coverage'] = round(
                case_stats[cid]['total_coverage'] / case_stats[cid]['sessions'])
    # 分类统计
    category_stats = {}
    for log_data in all_logs:
        for entry in log_data.get('log', []):
            cat = classify_consult_question(entry.get('student', ''))
            if cat not in category_stats:
                category_stats[cat] = {'count': 0, 'total': 0}
            category_stats[cat]['count'] += 1
    question_cats = {}
    for log_data in all_logs:
        for entry in log_data.get('log', []):
            q = entry.get('student', '')
            cat = classify_consult_question(q)
            if q not in question_cats:
                question_cats[q] = {'category': cat, 'count': 0, 'answers': []}
            question_cats[q]['count'] += 1
            if entry.get('patient'):
                question_cats[q]['answers'].append(entry['patient'][:80])
    return render_template('teacher.html',
                           logs=all_logs, case_stats=case_stats,
                           category_stats=category_stats, question_cats=question_cats,
                           total_sessions=len(all_logs))


def classify_consult_question(text):
    rules = [
        ('主诉-部位', ['位置', '部位', '哪里', '哪个牙']),
        ('主诉-时间', ['多久', '多长时间', '几天', '什么时候']),
        ('主诉-诱因', ['加重', '诱因', '什么情况', '怎样会']),
        ('现病史-疼痛', ['疼', '痛', '感觉', '性质', '怎么']),
        ('现病史-自发痛', ['自己', '自发', '夜间', '晚上', '睡觉']),
        ('现病史-用药', ['药', '吃过', '处理', '用']),
        ('既往史-全身病', ['病', '全身', '身体']),
        ('既往史-过敏', ['过敏']),
        ('既往史-牙科史', ['看过', '牙科', '以前', '治过']),
        ('生活习惯', ['抽烟', '喝酒', '吸烟']),
        ('基本信息', ['职业', '工作', '学校', '几岁', '多大']),
        ('外伤原因', ['受伤', '摔', '撞', '磕']),
        ('松动情况', ['松', '晃']),
    ]
    for cat, keywords in rules:
        if any(k in text for k in keywords):
            return cat
    return '其他'


@app.route('/api/consult/teacher/log/<consult_id>')
def consult_view_log_detail(consult_id):
    log_file = os.path.join(CONSULT_LOG_DIR, f'{consult_id}.json')
    if os.path.exists(log_file):
        with open(log_file, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))
    return jsonify({'error': '不存在'}), 404


@app.route('/analyze', methods=['POST'])
def analyze():
    """分析上传的照片"""
    if 'photos' not in request.files:
        return jsonify({'error': '未上传照片'}), 400

    files = request.files.getlist('photos')
    student_name = request.form.get('student_name', '匿名')
    student_class = request.form.get('student_class', '未知班级')
    tooth_type = request.form.get('tooth_type', 'real')
    operation_time = request.form.get('operation_time', None)

    if operation_time:
        try:
            operation_time = float(operation_time)
        except ValueError:
            operation_time = None

    # 保存上传的照片
    saved_paths = []
    session_id = uuid.uuid4().hex[:8]
    for i, file in enumerate(files):
        if file and allowed_file(file.filename):
            ext = file.filename.rsplit('.', 1)[1].lower()
            filename = f'{session_id}_{i}_{int(time.time())}.{ext}'
            filepath = UPLOAD_FOLDER / filename
            file.save(str(filepath))
            saved_paths.append(str(filepath))

    if not saved_paths:
        return jsonify({'error': '没有有效的照片文件（支持jpg/png/bmp）'}), 400

    # 分析每张照片，取最佳结果
    all_reports = []
    for path in saved_paths:
        report = engine.analyze(path, operation_time=operation_time, tooth_type=tooth_type)
        all_reports.append(report)

    # 选择总分最高的报告
    best_report = max(all_reports, key=lambda r: r.total_score)

    # 构建响应
    dimensions_data = []
    for d in best_report.dimensions:
        dimensions_data.append({
            'name': d.name,
            'score': d.score,
            'max_score': d.max_score,
            'percentage': round(d.score / d.max_score * 100, 1) if d.max_score > 0 else 0,
            'detail': d.detail,
            'status': d.status,
            'unit': d.unit,
            'raw_value': d.raw_value,
            'process_analysis': d.process_analysis,
            'targeted_suggestion': d.targeted_suggestion,
            'sub_scores': d.sub_scores if hasattr(d, 'sub_scores') else [],
        })

    result = {
        'session_id': session_id,
        'student_name': student_name,
        'student_class': student_class,
        'total_score': round(best_report.total_score, 1),
        'max_total': best_report.max_total,
        'percentage': round(best_report.total_score / best_report.max_total * 100, 1),
        'grade': ('优秀' if best_report.total_score >= 90 else
                  '良好' if best_report.total_score >= 80 else
                  '中等' if best_report.total_score >= 70 else
                  '及格' if best_report.total_score >= 60 else '不及格'),
        'dimensions': dimensions_data,
        'problem_areas': best_report.problem_areas,
        'suggestions': [d.targeted_suggestion for d in best_report.dimensions
                       if d.status in ('warning', 'bad') and d.targeted_suggestion][:5],
        'photo_count': len(saved_paths),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'overall_assessment': best_report.overall_assessment,
        'strengths': best_report.strengths,
        'weaknesses': best_report.weaknesses,
    }

    # 保存到历史记录
    scoring_history.append(result)
    if len(scoring_history) > 500:  # 保留最近500条
        scoring_history.pop(0)

    # 保存JSON报告
    report_json_path = REPORT_FOLDER / f'{session_id}.json'
    with open(report_json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return jsonify(result)


@app.route('/api/export')
def export_excel():
    """导出全班评分为Excel"""
    # 如果没有真实数据，加载所有保存的报告
    data_source = scoring_history if scoring_history else []
    if not data_source:
        # 尝试从reports文件夹加载
        for f in sorted(REPORT_FOLDER.glob('*.json')):
            try:
                with open(f, 'r', encoding='utf-8') as fh:
                    data_source.append(json.load(fh))
            except:
                pass

    if not data_source:
        return '暂无评分数据，请先进行至少一次评分后再导出。', 404

    wb = Workbook()

    # ── Sheet 1: 汇总表 ──
    ws1 = wb.active
    ws1.title = "评分汇总"

    # Style
    header_font = Font(name='微软雅黑', size=11, bold=True)
    title_font = Font(name='微软雅黑', size=14, bold=True)
    cell_font = Font(name='微软雅黑', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    header_font_white = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')

    # Title
    ws1.merge_cells('A1:P1')
    ws1.cell(row=1, column=1, value='II类洞制备AI评分汇总表').font = title_font
    ws1.cell(row=1, column=1).alignment = center

    # Headers
    headers = ['序号', '学生姓名', '班级', '总分', '等级', '外形轮廓(20)', '鸠尾峡(15)', '洞深(10)',
               '髓壁平面度(8)', '侧壁垂直度(7)', '点线角锐度(10)', '邻面盒形(10)', '洞缘光滑度(5)',
               '邻牙保护(10)', '操作过程(5)', '分析时间']

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data
    sorted_data = sorted(data_source, key=lambda x: x.get('total_score', 0), reverse=True)
    for i, d in enumerate(sorted_data, 1):
        row = i + 2
        ws1.cell(row=row, column=1, value=i).font = cell_font
        ws1.cell(row=row, column=1).alignment = center
        ws1.cell(row=row, column=2, value=d.get('student_name', '')).font = cell_font
        ws1.cell(row=row, column=2).alignment = center
        ws1.cell(row=row, column=3, value=d.get('student_class', '')).font = cell_font
        ws1.cell(row=row, column=3).alignment = center
        ws1.cell(row=row, column=4, value=d.get('total_score', 0)).font = Font(name='微软雅黑', size=10, bold=True)
        ws1.cell(row=row, column=4).alignment = center
        ws1.cell(row=row, column=5, value=d.get('grade', '')).font = cell_font
        ws1.cell(row=row, column=5).alignment = center

        # Dimension scores
        dims = {x['name']: x for x in d.get('dimensions', [])}
        dim_order = ['外形轮廓', '鸠尾峡比例', '洞深', '髓壁平面度', '侧壁垂直度',
                     '点线角锐度', '邻面盒形', '洞缘光滑度', '邻牙保护', '操作过程']
        for j, dim_name in enumerate(dim_order):
            dim = dims.get(dim_name, {})
            score = dim.get('score', '')
            max_s = dim.get('max_score', '')
            cell = ws1.cell(row=row, column=6+j, value=f'{score}/{max_s}' if score != '' else '')
            cell.font = cell_font
            cell.alignment = center

        ws1.cell(row=row, column=16, value=d.get('timestamp', '')).font = cell_font
        ws1.cell(row=row, column=16).alignment = center

        for col in range(1, 17):
            ws1.cell(row=row, column=col).border = thin_border

        # 等级着色
        grade_colors = {'优秀': '27ae60', '良好': '2B7BEC', '中等': 'f39c12', '及格': 'e67e22', '不及格': 'c0392b'}
        g = d.get('grade', '')
        if g in grade_colors:
            ws1.cell(row=row, column=5).font = Font(name='微软雅黑', size=10, bold=True, color=grade_colors[g])

    # 统计行
    stat_row = len(sorted_data) + 3
    ws1.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=3)
    ws1.cell(row=stat_row, column=1, value=f'合计: {len(sorted_data)}人').font = Font(name='微软雅黑', size=11, bold=True)
    ws1.cell(row=stat_row, column=1).alignment = center
    scores = [d.get('total_score', 0) for d in sorted_data]
    ws1.cell(row=stat_row, column=4, value=f'平均: {sum(scores)/len(scores):.1f}' if scores else 'N/A').font = Font(name='微软雅黑', size=11, bold=True)
    ws1.cell(row=stat_row, column=4).alignment = center
    pass_count = sum(1 for s in scores if s >= 60)
    ws1.cell(row=stat_row, column=5, value=f'及格率: {pass_count/len(scores)*100:.1f}%' if scores else 'N/A').font = Font(name='微软雅黑', size=11, bold=True)
    ws1.cell(row=stat_row, column=5).alignment = center
    for col in range(1, 17):
        ws1.cell(row=stat_row, column=col).border = thin_border

    # ── Sheet 2: 详细报告 ──
    ws2 = wb.create_sheet("详细分析")
    for i, d in enumerate(sorted_data[:30]):  # 最多30人详细
        row_start = i * 20 + 1
        # 学生信息
        ws2.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=6)
        ws2.cell(row=row_start, column=1,
                 value=f'{d.get("student_name","")} | {d.get("student_class","")} | 总分: {d.get("total_score",0)} | {d.get("grade","")}').font = Font(name='微软雅黑', size=13, bold=True)

        dims = {x['name']: x for x in d.get('dimensions', [])}
        dim_order = ['外形轮廓', '鸠尾峡比例', '洞深', '髓壁平面度', '侧壁垂直度',
                     '点线角锐度', '邻面盒形', '洞缘光滑度', '邻牙保护', '操作过程']

        for j, dim_name in enumerate(dim_order):
            row_d = row_start + 1 + j
            dim = dims.get(dim_name, {})
            ws2.cell(row=row_d, column=1, value=dim_name).font = Font(name='微软雅黑', size=10, bold=True)
            ws2.cell(row=row_d, column=2, value=f'{dim.get("score","")}/{dim.get("max_score","")}').font = cell_font
            detail = dim.get('detail', '')
            suggestion = dim.get('targeted_suggestion', '')
            ws2.cell(row=row_d, column=3, value=detail).font = cell_font
            ws2.cell(row=row_d, column=4, value=suggestion).font = cell_font

        # 整体评估
        row_d = row_start + 12
        ws2.cell(row=row_d, column=1, value='整体评估').font = Font(name='微软雅黑', size=10, bold=True)
        ws2.cell(row=row_d, column=2, value=d.get('overall_assessment','')).font = cell_font
        ws2.cell(row=row_d + 1, column=1, value='强项').font = Font(name='微软雅黑', size=10, bold=True)
        ws2.cell(row=row_d + 1, column=2, value=' | '.join(d.get('strengths', [])[:3])).font = cell_font
        ws2.cell(row=row_d + 2, column=1, value='弱项').font = Font(name='微软雅黑', size=10, bold=True)
        ws2.cell(row=row_d + 2, column=2, value=' | '.join(d.get('weaknesses', [])[:3])).font = cell_font

    # 列宽
    ws1.column_dimensions['A'].width = 6
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14
    for col in range(4, 17):
        ws1.column_dimensions[chr(64+col)].width = 13

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'II类洞评分汇总_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/report/<session_id>')
def view_report(session_id):
    """查看评分报告"""
    report_path = REPORT_FOLDER / f'{session_id}.json'
    if report_path.exists():
        with open(report_path, 'r', encoding='utf-8') as f:
            report = json.load(f)
        return render_template('report.html', report=report)
    return '报告不存在', 404


@app.route('/dashboard')
def dashboard():
    """教师端 — 全班统计分析"""
    if not scoring_history:
        # 生成模拟数据用于演示
        import random
        random.seed(42)
        demo_data = []
        for i in range(59):
            base = random.gauss(73, 12)
            base = max(30, min(98, base))
            demo_data.append({
                'student_name': f'学生{i+1:02d}',
                'student_class': '25口腔1班',
                'total_score': round(base, 1),
                'percentage': round(base, 1),
                'grade': ('优秀' if base >= 90 else '良好' if base >= 80 else
                          '中等' if base >= 70 else '及格' if base >= 60 else '不及格'),
                'dimensions': [
                    {'name': '外形轮廓', 'score': round(random.gauss(16, 3), 1), 'max_score': 20},
                    {'name': '鸠尾峡比例', 'score': round(random.gauss(11, 3), 1), 'max_score': 15},
                    {'name': '洞深', 'score': round(random.gauss(7, 2), 1), 'max_score': 10},
                    {'name': '髓壁平面度', 'score': round(random.gauss(5, 2.5), 1), 'max_score': 8},
                    {'name': '侧壁垂直度', 'score': round(random.gauss(4.5, 2), 1), 'max_score': 7},
                    {'name': '点线角锐度', 'score': round(random.gauss(7, 2), 1), 'max_score': 10},
                    {'name': '邻面盒形', 'score': round(random.gauss(7.5, 1.8), 1), 'max_score': 10},
                    {'name': '洞缘光滑度', 'score': round(random.gauss(4, 0.8), 1), 'max_score': 5},
                    {'name': '邻牙保护', 'score': round(random.gauss(8.5, 1.5), 1), 'max_score': 10},
                    {'name': '操作过程', 'score': round(random.gauss(3.5, 1), 1), 'max_score': 5},
                ],
            })
        history = demo_data
    else:
        history = scoring_history

    # 统计分析
    scores = [h['total_score'] for h in history]
    avg_score = sum(scores) / len(scores) if scores else 0
    pass_rate = sum(1 for s in scores if s >= 60) / len(scores) * 100 if scores else 0

    # 分数分布
    distribution = {'<60': 0, '60-69': 0, '70-79': 0, '80-89': 0, '90-100': 0}
    for s in scores:
        if s < 60: distribution['<60'] += 1
        elif s < 70: distribution['60-69'] += 1
        elif s < 80: distribution['70-79'] += 1
        elif s < 90: distribution['80-89'] += 1
        else: distribution['90-100'] += 1

    # 维度薄弱点分析
    dim_avg_scores = {}
    dim_names = list(SCORING_CONFIG.keys())
    dim_labels = {k: v['desc'] for k, v in SCORING_CONFIG.items()}
    for h in history:
        for d in h.get('dimensions', []):
            name = d['name']
            if name not in dim_avg_scores:
                dim_avg_scores[name] = {'total': 0, 'max': 0, 'count': 0}
            dim_avg_scores[name]['total'] += d['score']
            dim_avg_scores[name]['max'] += d['max_score']
            dim_avg_scores[name]['count'] += 1

    dim_analysis = []
    for name, data in dim_avg_scores.items():
        if data['count'] > 0:
            loss_rate = round((1 - data['total'] / data['max']) * 100, 1)
            dim_analysis.append({
                'name': name,
                'avg_score': round(data['total'] / data['count'], 1),
                'max_score': round(data['max'] / data['count'], 1),
                'loss_rate': loss_rate,
            })
    dim_analysis.sort(key=lambda x: x['loss_rate'], reverse=True)

    return render_template('dashboard.html',
                           history=history,
                           avg_score=round(avg_score, 1),
                           pass_rate=round(pass_rate, 1),
                           distribution=distribution,
                           dim_analysis=dim_analysis,
                           total_students=len(history),
                           config=SCORING_CONFIG)


@app.route('/api/history')
def api_history():
    """API: 获取评分历史"""
    return jsonify(scoring_history[-20:])  # 最近20条


@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)


# ── Main ──
@app.route('/analyze_endo', methods=['POST'])
def analyze_endo():
    """开髓洞形AI分析"""
    if 'photos' not in request.files:
        return jsonify({'error': '未上传照片'}), 400
    files = request.files.getlist('photos')
    saved_paths, session_id = [], uuid.uuid4().hex[:8]
    for i, file in enumerate(files):
        if file and allowed_file(file.filename):
            ext = file.filename.rsplit('.', 1)[1].lower()
            fp = UPLOAD_FOLDER / f'endo_{session_id}_{i}_{int(time.time())}.{ext}'
            file.save(str(fp)); saved_paths.append(str(fp))
    if not saved_paths: return jsonify({'error': '无有效文件'}), 400
    reports = [endo_engine.analyze(p) for p in saved_paths]
    best = max(reports, key=lambda r: r.total_score)
    dims = [{'name': d.name, 'score': d.score, 'max_score': d.max_score,
             'percentage': round(d.score/d.max_score*100,1) if d.max_score>0 else 0,
             'detail': d.detail, 'status': d.status,
             'process_analysis': d.process_analysis, 'targeted_suggestion': d.targeted_suggestion}
            for d in best.dimensions]
    result = {'session_id': session_id, 'total_score': round(best.total_score,1),
              'max_total': best.max_total, 'dimensions': dims,
              'grade': ('优秀' if best.total_score>=90 else '良好' if best.total_score>=80 else
                        '中等' if best.total_score>=70 else '及格' if best.total_score>=60 else '不及格'),
              'suggestions': [d.targeted_suggestion for d in best.dimensions
                             if d.status in ('warning','bad') and d.targeted_suggestion][:5],
              'photo_count': len(saved_paths),
              'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              'overall_assessment': best.overall_assessment,
              'strengths': best.strengths, 'weaknesses': best.weaknesses}
    with open(REPORT_FOLDER / f'endo_{session_id}.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    return jsonify(result)


@app.route('/analyze_xray', methods=['POST'])
def analyze_xray():
    """根管X光片AI分析"""
    if 'photos' not in request.files:
        return jsonify({'error': '未上传X光片'}), 400
    files = request.files.getlist('photos')
    saved_paths, session_id = [], uuid.uuid4().hex[:8]
    for i, file in enumerate(files):
        if file and allowed_file(file.filename):
            ext = file.filename.rsplit('.', 1)[1].lower()
            fp = UPLOAD_FOLDER / f'xray_{session_id}_{i}_{int(time.time())}.{ext}'
            file.save(str(fp)); saved_paths.append(str(fp))
    if not saved_paths: return jsonify({'error': '无有效文件'}), 400
    reports = [xray_engine.analyze(p) for p in saved_paths]
    best = max(reports, key=lambda r: r.total_score)
    dims = [{'name': d.name, 'score': d.score, 'max_score': d.max_score,
             'percentage': round(d.score/d.max_score*100,1) if d.max_score>0 else 0,
             'detail': d.detail, 'status': d.status,
             'process_analysis': d.process_analysis, 'targeted_suggestion': d.targeted_suggestion}
            for d in best.dimensions]
    result = {'session_id': session_id, 'total_score': round(best.total_score,1),
              'max_total': best.max_total, 'dimensions': dims,
              'grade': ('优秀' if best.total_score>=90 else '良好' if best.total_score>=80 else
                        '中等' if best.total_score>=70 else '及格' if best.total_score>=60 else '不及格'),
              'suggestions': [d.targeted_suggestion for d in best.dimensions
                             if d.status in ('warning','bad') and d.targeted_suggestion][:5],
              'photo_count': len(saved_paths),
              'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
              'overall_assessment': best.overall_assessment,
              'strengths': best.strengths, 'weaknesses': best.weaknesses}
    with open(REPORT_FOLDER / f'xray_{session_id}.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    return jsonify(result)


if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5050))
    print('=' * 60)
    print('  II类洞AI评分系统 v1.0')
    print('  学生端: http://localhost:5050/')
    print('  教师端: http://localhost:5050/dashboard')
    print('=' * 60)
    app.run(debug=False, host='0.0.0.0', port=port)
